"""Excel output, including the interactive Team Builder page.

Formulas are restricted to INDEX/MATCH and friends. FILTER/XLOOKUP/SORT are
"future functions" that need _xlfn./_xlfn._xlws. prefixes when a file is
written outside Excel and render as #NAME? if that is wrong; INDEX/MATCH also
works in older Excel and in Google Sheets.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def autosize(worksheet, max_width: int = 46) -> None:
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        longest = max((len(str(c.value)) for c in column_cells[:400] if c.value is not None),
                      default=8)
        worksheet.column_dimensions[letter].width = min(max(longest + 2, 10), max_width)


def style_sheet(worksheet) -> None:
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize(worksheet)


TITLE_FONT = Font(bold=True, size=14, color="1F3864")
SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="1F3864")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FONT = Font(italic=True, size=9, color="595959")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FINDER_SLOTS = 10
FINDER_HEADERS = ["Your partner's commander(s)", "Mode", "Mechanic", "Their colors",
                  "Their color identity", "Team colors", "Team identity", "# Shared sets",
                  "Shared set codes", "Shared sets", "Decks with their commander"]


def build_finder_sheet(book, n_options: int, n_configs: int, n_aliases: int,
                       n_partner_legal: int, n_commanders: int, capacity: int) -> None:
    """The input page: enter your own picks, get everything your partner may pick.

    Deliberately avoids FILTER/SORT/XLOOKUP. Those are "future functions" that
    must be written as _xlfn._xlws.FILTER etc. when a file is produced outside
    Excel, and resolve to #NAME? if that is wrong. INDEX/MATCH is portable and
    also works in older Excel and LibreOffice.
    """
    ws = book.create_sheet("Team Builder")
    opts = "'Partner Options'"       # quoted: these sheet names contain a space
    plegal = "'Partner Legal'"
    last_o = n_options + 1           # each helper table occupies rows 2..last_*
    last_l = n_configs + 1
    last_a = n_aliases + 1
    last_p = n_partner_legal + 1
    last_c = n_commanders + 1

    YOU1, YOU2, STATUS = 7, 8, 10     # input cells and the verdict line
    INFO1, INFO2 = 12, 13
    HINT, HEAD = 15, 16
    FIRST = HEAD + 1
    GRID_HEAD, GRID_FIRST = 5, 6      # collection grid, top right
    GRID_LAST = GRID_FIRST + FINDER_SLOTS - 1

    # $M$1 resolves the two typed cards to a canonical configuration label; every
    # other formula keys off it. The Aliases table carries both card orders, so
    # this never depends on Excel's text collation matching Python's.
    typed = f'IF($B${YOU2}="",$B${YOU1},$B${YOU1}&" + "&$B${YOU2})'

    def config_lookup(column: str) -> str:
        return (f'=IF($M$1="","",IFERROR(INDEX(Lists!${column}$2:${column}${last_l},'
                f'MATCH($M$1,Lists!$A$2:$A${last_l},0)),""))')

    ws["A1"] = "Two-Headed Giant Commander — Team Builder"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("You and your partner are teammates. Fill the yellow cells with your own picks; "
                "the table below lists every commander your partner may legally bring.")
    ws["A2"].font = NOTE_FONT

    # ---- Your picks ----
    ws["A4"] = "1. Your picks"
    ws["A4"].font = SECTION_FONT
    ws["A4"].fill = SECTION_FILL
    ws["A5"] = ("Leave the second cell blank unless you are running a partner pair "
                "(Partner, Partner with, Character select, Doctor's companion, Background).")
    ws["A5"].font = NOTE_FONT

    for row, label in ((YOU1, "Your commander:"),
                       (YOU2, "Your second commander (optional):")):
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        cell = ws[f"B{row}"]
        cell.fill = INPUT_FILL
        cell.border = BOX

    # Three distinguishable verdicts: legal, legal partners but never co-printed,
    # and not a partner combination at all.
    ws[f"A{STATUS}"] = (
        f'=IF($B${YOU1}="","Pick your commander above to begin.",'
        f'IF(AND($B${YOU2}<>"",$B${YOU1}=$B${YOU2}),"Pick two different cards.",'
        f'IF($M$1<>"",'
        f'IF($B${YOU2}="","OK — single commander. "&$M$3&" choices for your partner.",'
        f'"OK — legal partner pair ("&$M$4&"). "&$M$3&" choices for your partner."),'
        f'IF($B${YOU2}="",'
        f'IF($M$5="No (partner only)",'
        f'"This is a Background: it can only ever be a SECOND commander. Add a partner above.",'
        f'"That card is not an eligible commander in any precon deck."),'
        f'IFERROR(IF(MATCH($B${YOU1}&"|"&$B${YOU2},{plegal}!$A$2:$A${last_p},0)>0,'
        f'"These two legally partner, but they never appear together in one precon set."),'
        f'"These two cards cannot be your commanders together.")))))'
    )
    ws[f"A{STATUS}"].font = Font(bold=True, size=11)

    for col, label, source in (("A", "Your colors", "D"), ("C", "Color identity", "E"),
                               ("E", "Seat", "B"), ("G", "Partner mechanic", "C")):
        ws[f"{col}{INFO1}"] = label
        ws[f"{col}{INFO1}"].font = Font(bold=True, size=9)
        ws[f"{get_column_letter(ws[f'{col}{INFO1}'].column + 1)}{INFO1}"] = config_lookup(source)

    ws[f"A{INFO2}"] = "Locked into precon set(s)"
    ws[f"A{INFO2}"].font = Font(bold=True, size=9)
    ws[f"B{INFO2}"] = config_lookup("G")
    ws[f"E{INFO2}"] = "Choices for your partner"
    ws[f"E{INFO2}"].font = Font(bold=True, size=9)
    ws[f"F{INFO2}"] = "=$M$3"
    ws[f"G{INFO2}"] = "of which partner pairs"
    ws[f"G{INFO2}"].font = Font(bold=True, size=9)
    ws[f"H{INFO2}"] = (f'=IF($M$1="",0,COUNTIFS({opts}!$C$2:$C${last_o},$M$1,'
                       f'{opts}!$E$2:$E${last_o},"Partner"))')

    ws[f"A{HINT}"] = (
        f'=IF($M$1="","",IF($M$3=0,'
        f'"No legal choice for your partner: every other commander in your set overlaps your colors.",'
        f'"Your partner may pick any row below."))'
    )
    ws[f"A{HINT}"].font = NOTE_FONT

    for j, title in enumerate(FINDER_HEADERS):
        cell = ws[f"{get_column_letter(1 + j)}{HEAD}"]
        cell.value = title
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    # One MATCH locates the block; every row then reads a fixed offset from it.
    MECHANIC_COL = "F"   # Partner Options column F is blank for single commanders
    for k in range(capacity):
        row = FIRST + k
        for j in range(len(FINDER_HEADERS)):
            src = get_column_letter(4 + j)   # Partner Options column D..N
            lookup = f"INDEX({opts}!${src}$2:${src}${last_o},$M$2+ROW()-{FIRST})"
            # A formula pointing at an empty cell renders as 0, not blank, so the
            # one column that is legitimately empty has to be coerced back.
            value = f'IF({lookup}="","",{lookup})' if src == MECHANIC_COL else lookup
            ws[f"{get_column_letter(1 + j)}{row}"] = (
                f'=IF(OR($M$2=0,ROW()-{HEAD}>$M$3),"",{value})'
            )

    # ---- Collection grid, kept top-right so it stays inside the frozen pane ----
    ws[f"O2"] = "2. Which commanders in your collection can team up"
    ws["O2"].font = SECTION_FONT
    ws["O2"].fill = SECTION_FILL
    ws["O3"] = (f"Enter up to {FINDER_SLOTS} single commanders. Each cell shows the precon set "
                "code(s) that make those two a legal team; “—” means they cannot.")
    ws["O3"].font = NOTE_FONT
    ws[f"O{GRID_HEAD}"] = "Your collection"
    ws[f"O{GRID_HEAD}"].font = Font(bold=True)
    for j in range(FINDER_SLOTS):
        col = get_column_letter(16 + j)
        header = ws[f"{col}{GRID_HEAD}"]
        header.value = f'=IF($O${GRID_FIRST + j}="","",$O${GRID_FIRST + j})'
        header.font = Font(bold=True, size=8)
        header.alignment = Alignment(wrap_text=True, vertical="bottom")

    for i in range(FINDER_SLOTS):
        row = GRID_FIRST + i
        ws[f"O{row}"].fill = INPUT_FILL
        ws[f"O{row}"].border = BOX
        for j in range(FINDER_SLOTS):
            col = get_column_letter(16 + j)
            target = ws[f"{col}{row}"]
            target.value = (
                f'=IF(OR($O{row}="",{col}${GRID_HEAD}=""),"",'
                f'IF($O{row}={col}${GRID_HEAD},"",'
                f'IFERROR(INDEX({opts}!$L$2:$L${last_o},'
                f'MATCH($O{row}&"|"&{col}${GRID_HEAD},{opts}!$A$2:$A${last_o},0)),"—")))'
            )
            target.alignment = Alignment(horizontal="center")
            target.border = BOX
            target.font = Font(size=8)

    # Hidden helpers.
    ws["M1"] = (f'=IF($B${YOU1}="","",IFERROR(INDEX(Aliases!$B$2:$B${last_a},'
                f'MATCH({typed},Aliases!$A$2:$A${last_a},0)),""))')
    ws["M2"] = f'=IF($M$1="",0,IFERROR(MATCH($M$1&"#1",{opts}!$B$2:$B${last_o},0),0))'
    ws["M3"] = f'=IF($M$1="",0,COUNTIF({opts}!$C$2:$C${last_o},$M$1))'
    ws["M4"] = (f'=IF($M$1="","",IFERROR(INDEX(Lists!$C$2:$C${last_l},'
                f'MATCH($M$1,Lists!$A$2:$A${last_l},0)),""))')
    ws["M5"] = (f'=IF($B${YOU1}="","",IFERROR(INDEX(Commanders!$E$2:$E${last_c},'
                f'MATCH($B${YOU1},Commanders!$A$2:$A${last_c},0)),""))')
    ws.column_dimensions["M"].hidden = True

    # showDropDown is inverted in the file format: True hides the in-cell arrow.
    # Both pickers list individual cards, not pre-combined pairs.
    validation = DataValidation(type="list", formula1=f"Commanders!$A$2:$A${last_c}",
                                allow_blank=True, showDropDown=False,
                                errorTitle="Not a known commander",
                                error="Pick a value from the dropdown, or copy a name exactly "
                                      "as it appears on the Commanders sheet.")
    ws.add_data_validation(validation)
    validation.add(f"B{YOU1}:B{YOU2}")
    validation.add(f"O{GRID_FIRST}:O{GRID_LAST}")

    ws.column_dimensions["A"].width = 40
    for j in range(1, len(FINDER_HEADERS)):
        ws.column_dimensions[get_column_letter(1 + j)].width = 20
    ws.column_dimensions["O"].width = 34
    for j in range(FINDER_SLOTS):
        ws.column_dimensions[get_column_letter(16 + j)].width = 11
    ws.freeze_panes = f"A{FIRST}"


def write_workbook(path: Path, frames: dict[str, pd.DataFrame],
                   finder: dict | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        path.open("ab").close()
    except PermissionError:
        raise SystemExit(
            f"Cannot write {path} -- it is open in Excel. Close it and re-run, "
            f"or pass --out with a different path."
        ) from None
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in frames.items():
            frame.to_excel(writer, sheet_name=name, index=False)
        book = writer.book
        for name in frames:
            if name != "Rules":
                style_sheet(book[name])
        rules = book["Rules"]
        for cell in rules[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        rules.column_dimensions["A"].width = 28
        rules.column_dimensions["B"].width = 110
        for row in rules.iter_rows(min_row=2, min_col=2, max_col=2):
            row[0].alignment = Alignment(wrap_text=True, vertical="top")

        if finder:
            build_finder_sheet(book, **finder)
            sheet = book["Team Builder"]
            book.move_sheet(sheet, offset=-(book.index(sheet) - 1))
            for helper in ("Lists", "Aliases", "Partner Legal"):
                book[helper].sheet_state = "hidden"
            book.active = 1
